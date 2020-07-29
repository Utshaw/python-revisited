#!/usr/bin/env python3
# This example demonstrates how to open xlsx file & work with it
# Spreadsheet document is called workbook
# The workbook can contain multiple Sheets/Worksheets
import openpyxl
import os

workbook = openpyxl.load_workbook('sample-spreadsheet.xlsx')
print(workbook.sheetnames)  # list of all worksheets in this workbook
sheet = workbook['Sheet1'] 
cell = sheet['G1'] # select a cell
print(str(cell.value) + " - " + str(type(cell)))
cell2 = sheet.cell(row=1, column=7)
print(cell2.value)

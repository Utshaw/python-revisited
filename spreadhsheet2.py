#!/usr/bin/env python3
# This example demonstrates how to create xlsx file using python
# Spreadsheet document is called workbook
# The workbook can contain multiple Sheets/Worksheets
import openpyxl
import os

workbook = openpyxl.Workbook()
print(workbook.sheetnames) # starts with a single sheet named "Sheet"
sheet = workbook['Sheet']
sheet['A1'] = 'Name'
sheet['B1'] = 'Roll'
sheet['A2'] = 'Farhan Utshaw'
sheet['B2'] = '1505105'
sheet2 = workbook.create_sheet(index=0) # if index is given this new sheet will be placed at that index
sheet2.title="My new Sheet"
print("After creating new sheet: " +  str(workbook.sheetnames))
workbook.save('generated-sample.xlsx')